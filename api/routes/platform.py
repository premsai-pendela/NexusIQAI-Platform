"""NexusIQAI Platform Mode routes — multi-company, role-aware analyst API.

Every data route depends on AccessContext resolved from the session token.
Company and role always come from the server-side registry; no client-sent
company/role value is ever trusted.
"""

import asyncio
import json
import time
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field

from api.serializers import build_answer_payload
from nexus_platform import store
from nexus_platform.auth import (AccessContext, create_token,
                                 get_access_context, require_admin)
from nexus_platform.brain_builder import brain_status, build_brain
from nexus_platform.contexts import brain_dir
from nexus_platform.registry import get_registry

router = APIRouter(prefix="/platform")

_QUERY_TIMEOUT = 90


# ── Schemas ─────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    email: str = Field(..., max_length=120)
    password: str = Field(..., max_length=120)


class PlatformQueryRequest(BaseModel):
    question: str = Field(..., min_length=3, max_length=500)
    session_id: str = Field(..., min_length=4, max_length=64)
    repeat_action: Optional[str] = Field(
        None, pattern="^(use_previous|rerun|analyze_with_ai)$")
    # A caller (e.g. a simulation employee) may opt its own traffic into the
    # "simulated" bucket. Safe by construction: "simulated" is filtered OUT of
    # default reads, so a client can only make its own traffic MORE hidden,
    # never forge "real" traffic. Anything else falls back to real.
    source: Optional[str] = Field(None, pattern="^(real|simulated)$")


class FeedbackRequest(BaseModel):
    category: str = Field(..., max_length=40)
    message: str = Field(..., min_length=3, max_length=2000)
    page: Optional[str] = Field(None, max_length=120)
    trace_id: Optional[str] = Field(None, max_length=40)


class FeedbackStatusRequest(BaseModel):
    status: str = Field(..., pattern="^(new|reviewed|resolved)$")


class XlsxExportRequest(BaseModel):
    title: str = Field(..., min_length=1, max_length=80)
    rows: list[dict] = Field(..., min_length=1, max_length=5000)
    question: Optional[str] = Field(None, max_length=500)
    trace_id: Optional[str] = Field(None, max_length=40)


def _profile(ctx: AccessContext) -> dict:
    return {
        "email": ctx.employee.email,
        "name": ctx.employee.name,
        "title": ctx.employee.title,
        "role": ctx.employee.role,
        "is_admin": ctx.is_admin,
        "company": {
            "slug": ctx.company.slug,
            "name": ctx.company.name,
            "industry": ctx.company.industry,
            "description": ctx.company.description,
        },
        "access": {
            "summary": ctx.policy.summary,
            "denied_summary": ctx.policy.denied_summary,
            "tables": list(ctx.policy.allowed_tables),
            "departments": list(ctx.policy.allowed_departments),
            "read_only": True,
        },
    }


# ── Auth ────────────────────────────────────────────────────────────────

@router.post("/login")
def login(req: LoginRequest):
    employee = get_registry().authenticate(req.email, req.password)
    if employee is None:
        raise HTTPException(status_code=401, detail="Invalid work email or password")
    token = create_token(employee.email)
    from nexus_platform.access_policy import get_policy
    registry = get_registry()
    ctx = AccessContext(employee=employee,
                        company=registry.get_company(employee.company_slug),
                        policy=get_policy(employee.role))
    return {"token": token, "profile": _profile(ctx)}


@router.get("/me")
def me(ctx: AccessContext = Depends(get_access_context)):
    return _profile(ctx)


# ── Workspace / brain ───────────────────────────────────────────────────

@router.get("/workspace")
def workspace(ctx: AccessContext = Depends(get_access_context)):
    status = brain_status(ctx.company.slug)
    if not ctx.is_admin:
        # Employees see readiness only — not rebuild internals or changed files
        status = {"status": "ready" if status["status"] != "not_built" else "not_built",
                  "built_at": status.get("built_at")}
    schema = {}
    catalog_path = brain_dir(ctx.company.slug) / "schema_catalog.json"
    if catalog_path.exists():
        full = json.loads(catalog_path.read_text())
        schema = {t: v for t, v in full.items() if t in ctx.policy.allowed_tables}
    docs = []
    inventory_path = brain_dir(ctx.company.slug) / "doc_inventory.json"
    if inventory_path.exists():
        docs = [d for d in json.loads(inventory_path.read_text())
                if d.get("department") in ctx.policy.allowed_departments]
    return {
        "profile": _profile(ctx),
        "brain": status,
        "tables": schema,
        "documents": docs,
    }


@router.post("/brain/rebuild")
def rebuild_brain(ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    log = build_brain(ctx.company.slug)
    # Rebuilt indexes are picked up by RAG agents via the ingestion version file.
    return {"status": "ok", "build_log": log}


@router.get("/brain/status")
def get_brain_status(ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    return brain_status(ctx.company.slug)


# ── Query ───────────────────────────────────────────────────────────────

@router.post("/query")
async def platform_query(req: PlatformQueryRequest,
                         ctx: AccessContext = Depends(get_access_context)):
    from nexus_platform.query_service import run_query
    request_id = str(uuid.uuid4())[:8]
    start = time.time()
    loop = asyncio.get_event_loop()
    src = req.source if req.source in ("real", "simulated") else "real"

    def _run():
        # Set the source tag inside the executor thread so save_trace (which
        # runs here) sees it — a contextvar set on the async thread would not
        # propagate into the executor.
        with store.tagged_trace_source(src):
            return run_query(ctx, req.question, req.session_id,
                             repeat_action=req.repeat_action)

    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(None, _run),
            timeout=_QUERY_TIMEOUT,
        )
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="Query timed out")

    payload = build_answer_payload(result)
    platform_meta = result.get("platform") or {}
    sources = []
    for s in result.get("sources") or []:
        if isinstance(s, dict):
            sources.append({
                "type": s.get("type", "rag"),
                "content": str(s.get("content", s.get("document", "")))[:500],
                "filename": s.get("filename") or s.get("source"),
            })
    return {
        **payload,
        "sources": sources,
        "platform": platform_meta,
        "latency_ms": (time.time() - start) * 1000,
        "request_id": request_id,
    }


@router.post("/export/xlsx")
def export_xlsx(req: XlsxExportRequest,
                ctx: AccessContext = Depends(get_access_context)):
    """Turn chart/table rows the caller already received into an .xlsx file.

    Exports only data sent back by the client (already access-filtered when it
    was answered); the session dependency keeps the endpoint company-scoped.
    """
    import io
    import re as _re

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from datetime import datetime, timezone

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append([f"{ctx.company.name} — {req.title}"])
    ws["A1"].font = Font(bold=True, size=13)
    cols = list(req.rows[0].keys())
    ws.append(cols)
    for cell in ws[2]:
        cell.font = Font(bold=True)
    for row in req.rows:
        ws.append([row.get(c) for c in cols])
    for idx, col in enumerate(cols, start=1):
        width = max(len(str(col)), *(len(str(r.get(col, ""))) for r in req.rows[:200]))
        ws.column_dimensions[ws.cell(row=2, column=idx).column_letter].width = min(width + 2, 40)

    # Provenance sheet: who exported what, under which access scope
    meta = wb.create_sheet("Provenance")
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 70
    for label, value in [
        ("question", req.question or req.title),
        ("company", ctx.company.name),
        ("employee", ctx.employee.email),
        ("role", ctx.employee.role),
        ("access scope", ", ".join(ctx.policy.allowed_tables) or "none"),
        ("trace id", req.trace_id or "—"),
        ("exported at", datetime.now(timezone.utc).isoformat()),
        ("note", "Role-filtered workspace data — same rows shown in the NexusIQAI UI."),
    ]:
        meta.append([label, value])
        meta.cell(row=meta.max_row, column=1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = _re.sub(r"[^a-z0-9]+", "-", req.title.lower()).strip("-")[:40] or "export"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{slug}.xlsx"'},
    )


@router.get("/history")
def history(ctx: AccessContext = Depends(get_access_context)):
    """The caller's own recent question history (their memory, their company)."""
    return {"turns": store.employee_history(ctx.company.slug, ctx.employee.email, limit=30)}


# ── Feedback ────────────────────────────────────────────────────────────

@router.post("/feedback")
def submit_feedback(req: FeedbackRequest,
                    ctx: AccessContext = Depends(get_access_context)):
    fid = store.save_feedback(ctx.company.slug, ctx.employee.email,
                              ctx.employee.role, req.category, req.message,
                              req.page, req.trace_id)
    return {"status": "ok", "feedback_id": fid}


@router.get("/admin/feedback")
def review_feedback(employee: Optional[str] = None, status: Optional[str] = None,
                    category: Optional[str] = None,
                    ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    return {"feedback": store.list_feedback(ctx.company.slug, employee=employee,
                                            status=status, category=category)}


@router.patch("/admin/feedback/{feedback_id}")
def set_feedback_status(feedback_id: str, req: FeedbackStatusRequest,
                        ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    if not store.update_feedback_status(ctx.company.slug, feedback_id, req.status):
        raise HTTPException(status_code=404, detail="Feedback not found in your company workspace")
    return {"status": "ok"}


# ── Trace review (Admin/CEO, same company only) ─────────────────────────

@router.get("/admin/traces")
def review_traces(employee: Optional[str] = None, date_from: Optional[str] = None,
                  date_to: Optional[str] = None, source: Optional[str] = None,
                  ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    # source=None returns both real and simulated, each carrying its own
    # `source` badge — labelled, never conflated.
    return {"traces": store.list_traces_for_review(
        ctx.company.slug, employee=employee, date_from=date_from,
        date_to=date_to, source=source)}


@router.get("/traces/{trace_id}")
def get_trace(trace_id: str, ctx: AccessContext = Depends(get_access_context)):
    """Company-scoped trace detail. Employees may open only their own traces;
    Admin/CEO may open any trace in their company."""
    trace = store.get_trace(ctx.company.slug, trace_id)
    if trace is None:
        raise HTTPException(status_code=404, detail="Trace not found in your company workspace")
    if not ctx.is_admin and trace["employee"] != ctx.employee.email:
        raise HTTPException(status_code=403, detail="You can only view your own traces")
    # The answer text lives in the memory turn the query wrote; surface it so
    # the detail pane can show what the analyst actually answered.
    trace["answer"] = store.answer_for_trace(ctx.company.slug, trace_id)
    return trace


# ── Analyst Health Check (Admin/CEO, same company only) ─────────────────

class HealthCheckRequest(BaseModel):
    window_days: int = Field(30, ge=1, le=730)
    llm_summary: bool = False
    source: str = "real"  # "real" | "simulated" — audited traffic, never mixed


@router.post("/admin/health-check")
async def run_health_check_route(req: HealthCheckRequest,
                                 ctx: AccessContext = Depends(get_access_context)):
    """Run the Analyst Health Check agent over this company's traces and
    feedback. Analysis is deterministic; the optional LLM executive summary
    degrades honestly when providers are exhausted. `source` selects real
    traffic (default) or synthetic-demo traffic — the two are never mixed in
    one report."""
    require_admin(ctx)
    source = req.source if req.source in ("real", "simulated") else "real"
    from nexus_platform.health_check import run_health_check
    loop = asyncio.get_event_loop()
    report = await loop.run_in_executor(
        None, lambda: run_health_check(
            ctx.company.slug, requested_by=ctx.employee.email,
            window_days=req.window_days, llm_summary=req.llm_summary,
            source=source))
    return report


class HealthReviewRequest(BaseModel):
    window_days: int = Field(30, ge=1, le=730)
    source: str = "real"          # "real" | "simulated"
    llm_budget: int = Field(25, ge=0, le=200)


@router.post("/admin/health-review")
async def run_health_review_route(req: HealthReviewRequest,
                                  ctx: AccessContext = Depends(get_access_context)):
    """Wave 1 of the Health Check agent: grade every trace (deterministic →
    capped LLM → human-review) and return a structured, downloadable report."""
    require_admin(ctx)
    source = req.source if req.source in ("real", "simulated") else "real"
    from nexus_platform.health_review import run_health_review
    loop = asyncio.get_event_loop()
    report = await loop.run_in_executor(
        None, lambda: run_health_review(
            ctx.company.slug, requested_by=ctx.employee.email,
            window_days=req.window_days, source=source,
            llm_budget=req.llm_budget))
    return report


@router.get("/admin/health-reports")
def list_health_reports_route(ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    return {"reports": store.list_health_reports(ctx.company.slug)}


@router.get("/admin/health-reports/{report_id}")
def get_health_report_route(report_id: str,
                            ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    report = store.get_health_report(ctx.company.slug, report_id)
    if report is None:
        raise HTTPException(status_code=404,
                            detail="Report not found in your company workspace")
    return report


@router.get("/admin/employees")
def company_employees(ctx: AccessContext = Depends(get_access_context)):
    require_admin(ctx)
    return {"employees": [
        {"email": e.email, "name": e.name, "role": e.role, "title": e.title}
        for e in get_registry().company_employees(ctx.company.slug)
    ]}
