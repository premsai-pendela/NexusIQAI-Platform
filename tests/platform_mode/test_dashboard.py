"""Dashboard + export tests: role-filtered blocks, no cross-role leakage."""

import os

import pytest
from fastapi.testclient import TestClient

os.environ["NEXUSIQ_PREWARM_LIVE"] = "false"

from api.main import app  # noqa: E402
from nexus_platform.access_policy import get_policy  # noqa: E402
from nexus_platform.auth import AccessContext  # noqa: E402
from nexus_platform.dashboard import build_dashboard, wants_dashboard  # noqa: E402
from nexus_platform.registry import get_registry  # noqa: E402


def ctx_for(email: str) -> AccessContext:
    r = get_registry()
    e = r.get_employee(email)
    return AccessContext(employee=e, company=r.get_company(e.company_slug),
                         policy=get_policy(e.role))


def test_wants_dashboard_detection():
    assert wants_dashboard("Give me a dashboard")
    assert wants_dashboard("show a business overview please")
    assert not wants_dashboard("What was revenue in Q3?")


def test_analyst_dashboard_has_no_hr_blocks():
    d = build_dashboard(ctx_for("analyst@acmecloud.test"))
    assert d is not None and d["kpis"] and d["charts"]
    titles = [k["title"] for k in d["kpis"]] + [c["title"] for c in d["charts"]]
    assert not any("Headcount" in t or "attrition" in t.lower() or "termination" in t.lower()
                   for t in titles)
    assert all("employees_hr" not in sql for sql in d["sql_used"])


def test_hr_dashboard_is_hr_only():
    d = build_dashboard(ctx_for("hr@acmecloud.test"))
    assert d is not None
    for sql in d["sql_used"]:
        assert "employees_hr" in sql
        assert "orders" not in sql and "invoices" not in sql


def test_admin_dashboard_richer_than_hr():
    admin = build_dashboard(ctx_for("admin@acmecloud.test"))
    hr = build_dashboard(ctx_for("hr@acmecloud.test"))
    assert len(admin["sql_used"]) > len(hr["sql_used"])


def test_unknown_role_gets_no_dashboard():
    from nexus_platform.registry import Employee, hash_password
    r = get_registry()
    fake = Employee(email="x@acmecloud.test", name="X", company_slug="acmecloud",
                    role="Intern", password_hash=hash_password("x"), title="Intern")
    ctx = AccessContext(employee=fake, company=r.get_company("acmecloud"),
                        policy=get_policy("Intern"))
    assert build_dashboard(ctx) is None


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


def test_xlsx_export_roundtrip(client):
    token = client.post("/api/v1/platform/login",
                        json={"email": "analyst@acmecloud.test",
                              "password": "demo-analyst-2026"}).json()["token"]
    r = client.post("/api/v1/platform/export/xlsx",
                    headers={"X-NexusIQ-Session": token},
                    json={"title": "Revenue by region",
                          "rows": [{"region": "East", "revenue": 100.5},
                                   {"region": "West", "revenue": 90.25}]})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml")
    # openpyxl can read it back
    import io

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws["A2"].value == "region"
    assert ws["B3"].value == 100.5


def test_xlsx_export_requires_session(client):
    r = client.post("/api/v1/platform/export/xlsx",
                    json={"title": "x", "rows": [{"a": 1}]})
    assert r.status_code == 401
